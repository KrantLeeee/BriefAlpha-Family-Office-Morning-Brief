"""APScheduler cron jobs.

Per design.md §8 + tasks 12.x / 20.7.x:

| Cron                      | Action                                              |
|---------------------------|-----------------------------------------------------|
| weekday 06:50 HKT         | refresh sensitive_entity_dictionary                 |
| weekday 07:00 HKT         | rolling ingestion (warm-up, low pressure)           |
| weekday 07:55 HKT         | freeze evidence_pool + run_brief                    |
| daily 16:00 HKT           | purge alias_map ciphertext                          |
| daily 03:30 HKT           | retention sweep — PDFs older than 7d, audit > 90d   |
| every 5 minutes           | aggregate source_health snapshot to redis           |
| every minute              | research worker tick (drain queues)                 |
| Monday 04:00 HKT          | refresh SEC + HKEX symbol maps                      |

The schedule starts inside `main.lifespan` so a single-process MVP doesn't
need a separate cron daemon.
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from apscheduler.jobstores.sqlalchemy import SQLAlchemyJobStore
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

from briefalpha_api.anonymization import delete_alias_map
from briefalpha_api.audit import aggregate_source_health
from briefalpha_api.research import research_worker_tick, sweep_old_files
from briefalpha_api.settings import DATA_DIR

log = logging.getLogger("briefalpha.scheduler")
HKT_NAME = "Asia/Hong_Kong"
HKT = ZoneInfo(HKT_NAME)


# ---------------------------------------------------------------------------
# Job implementations
# ---------------------------------------------------------------------------


def _purge_alias_maps() -> None:
    folder: Path = DATA_DIR / "alias_maps"
    if not folder.exists():
        return
    for path in folder.glob("*.enc"):
        brief_id = path.stem
        try:
            delete_alias_map(brief_id)
        except Exception as exc:  # noqa: BLE001
            log.warning("alias_map purge failed %s: %s", brief_id, exc)


async def _rolling_ingestion_job() -> None:
    """07:00 warm-up: pull market / news / official sources into the
    raw_payload cache + record source_health rows. Does NOT freeze the
    evidence pool or generate the brief — that happens at 07:55.

    This separation gives operators a quiet ~55-minute window to spot
    source degradation (via /api/source-health) before the brief locks.
    """
    from briefalpha_api.db.session import SessionLocal
    from briefalpha_api.ingestion.runner import run_ingestion
    from briefalpha_api.portfolio.repo import load_positions, load_watchlist
    from briefalpha_api.portfolio.universe import build_universe

    today = datetime.now(tz=HKT).strftime("%Y-%m-%d")
    log.info("scheduler: 07:00 rolling ingestion warm-up for %s", today)
    try:
        async with SessionLocal() as s:
            positions = await load_positions(s, user_id="demo")
            watchlist = await load_watchlist(s, user_id="demo")
        if not positions:
            log.warning("rolling ingestion: no portfolio rows for demo user — skipping")
            return
        universe, _bucket_summary = build_universe(
            brief_id=f"warmup-{today}", positions=positions, watchlist=watchlist
        )
        result = await run_ingestion(universe)
        per_source = {name: len(items) for name, items in result.items()}
        log.info("scheduler: rolling ingestion ok %s", per_source)
    except Exception as exc:  # noqa: BLE001
        log.exception("scheduler: rolling ingestion failed: %s", exc)


async def _freeze_and_run_brief_job() -> None:
    """07:55 freeze + brief generation. Soft-fails (logs only) so a one-off
    pipeline error doesn't kill the scheduler — the brief route falls back
    to the prior day with `stale=True` if the cache stays cold."""
    from briefalpha_api.cache import set_brief_cache
    from briefalpha_api.pipeline.run import run_full_brief

    today = datetime.now(tz=HKT).strftime("%Y-%m-%d")
    log.info("scheduler: 07:55 freeze + run_brief for %s starting", today)
    try:
        artifact = await run_full_brief(today)
        await set_brief_cache(today, artifact)
        log.info("scheduler: 07:55 brief for %s ok", today)
    except Exception as exc:  # noqa: BLE001
        log.exception("scheduler: 07:55 brief for %s failed: %s", today, exc)


async def _refresh_dictionary() -> None:
    """Rebuild the sensitive_entity_dictionary cache.

    Today this only invalidates the lru_cache so the next call rebuilds
    from `company_alias_zh.yml`. Live yfinance lookups are deferred to the
    main pipeline path (which already owns universe construction).
    """
    from briefalpha_api.anonymization.sensitive_entity_dictionary import _load_zh_aliases

    _load_zh_aliases.cache_clear()
    log.info("scheduler: sensitive dict cache cleared")


async def _aggregate_source_health() -> None:
    try:
        await aggregate_source_health()
    except Exception as exc:  # noqa: BLE001
        log.warning("source_health aggregate failed: %s", exc)


async def _research_worker_tick() -> None:
    try:
        n = await research_worker_tick()
        if n:
            log.info("scheduler: research worker processed %d jobs", n)
    except Exception as exc:  # noqa: BLE001
        log.warning("research worker tick failed: %s", exc)


async def _retention_sweep() -> None:
    try:
        n_pdfs = sweep_old_files(max_age_days=7)
        if n_pdfs:
            log.info("retention sweep: removed %d encrypted PDFs", n_pdfs)
    except Exception as exc:  # noqa: BLE001
        log.warning("retention sweep failed: %s", exc)


def _transform_sec(content: bytes) -> bytes:
    """Transform raw SEC company_tickers.json into the loader's
    `{"mappings": {ticker: padded_cik}}` shape.

    Raw SEC schema: `{"0": {"cik_str": int, "ticker": str, "title": str}, ...}`.
    CIK is zero-padded to 10 chars to match the SEC EDGAR URL contract.
    """
    import json as _json

    data = _json.loads(content)
    mappings: dict[str, str] = {}
    for entry in data.values():
        if not isinstance(entry, dict):
            continue
        ticker = entry.get("ticker")
        cik = entry.get("cik_str")
        if not ticker or cik is None:
            continue
        mappings[str(ticker).upper()] = str(cik).zfill(10)
    return _json.dumps({"mappings": mappings}, ensure_ascii=False).encode("utf-8")


def _transform_hkex(content: bytes) -> bytes:
    """Transform HKEX ListOfSecurities.xlsx into `{"mappings": {ticker: code}}`.

    The xlsx has a 5-digit numeric stock code in column A; we coerce that into
    the canonical 4-digit-padded HK ticker (00700 → 0700.HK). 5-digit codes
    (warrants/derivatives at 60000+) keep their natural width. The rest of the
    codebase (portfolio fixtures, alias variants, treemap palettes) all use
    the 4-digit form, so anything else here would silently miss lookups.
    """
    import io as _io
    import json as _json

    from openpyxl import load_workbook

    # NOTE: read_only=True is intentionally NOT used. HKEX's ListOfSecurities.xlsx
    # ships an inaccurate worksheet `dimension` tag in its XML, and read_only mode
    # honors that tag literally — it stops after ~5 rows even though the file
    # actually contains ~17.8k. The full file is ~1.4 MB so the memory cost of
    # loading non-streaming once a week is negligible.
    wb = load_workbook(_io.BytesIO(content), data_only=True)
    ws = wb.active
    mappings: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        cell = row[0]
        if cell is None:
            continue
        s = str(cell).strip()
        if not s.isdigit() or len(s) > 5:
            continue
        code = s.zfill(5)
        stripped = code.lstrip("0") or "0"
        ticker_num = stripped if len(stripped) >= 4 else stripped.zfill(4)
        mappings[f"{ticker_num}.HK"] = code
    return _json.dumps({"mappings": mappings}, ensure_ascii=False).encode("utf-8")


async def _refresh_symbol_maps() -> None:
    """Best-effort weekly download + transform of public symbol maps.

    We tolerate network failures so cron noise doesn't accumulate offline.
    The on-disk files are the loader-friendly `{"mappings": {...}}` shape,
    NOT the raw upstream payload — the loader (`ingestion/symbol_map.py`)
    only knows how to read `mappings`.
    """
    import httpx

    targets = [
        (
            "sec_company_tickers.json",
            "https://www.sec.gov/files/company_tickers.json",
            _transform_sec,
        ),
        (
            "hkex_stock_codes.json",
            "https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/ListOfSecurities.xlsx",
            _transform_hkex,
        ),
    ]
    out_dir = DATA_DIR / "symbol_maps"
    out_dir.mkdir(parents=True, exist_ok=True)
    headers = {"User-Agent": "BriefAlpha/0.1 ops@briefalpha.local"}

    async with httpx.AsyncClient(timeout=30.0, headers=headers) as client:
        for fname, url, transform in targets:
            try:
                r = await client.get(url)
                r.raise_for_status()
                payload = transform(r.content)
                (out_dir / fname).write_bytes(payload)
                log.info("symbol_map refresh: %s ok (%d bytes)", fname, len(payload))
            except Exception as exc:  # noqa: BLE001
                log.warning("symbol_map refresh %s failed: %s", fname, exc)


# ---------------------------------------------------------------------------
# Scheduler builder
# ---------------------------------------------------------------------------


def build_scheduler(jobstore_url: str | None = None) -> AsyncIOScheduler:
    if jobstore_url is None:
        jobstore_url = f"sqlite:///{DATA_DIR / 'briefalpha.db'}"
    scheduler = AsyncIOScheduler(
        jobstores={"default": SQLAlchemyJobStore(url=jobstore_url)},
        timezone=HKT_NAME,
    )

    scheduler.add_job(
        _refresh_dictionary,
        trigger=CronTrigger(day_of_week="mon-fri", hour=6, minute=50),
        id="refresh_sensitive_dict",
        replace_existing=True,
    )
    scheduler.add_job(
        _rolling_ingestion_job,
        trigger=CronTrigger(day_of_week="mon-fri", hour=7, minute=0),
        id="rolling_ingestion",
        replace_existing=True,
    )
    scheduler.add_job(
        _freeze_and_run_brief_job,
        trigger=CronTrigger(day_of_week="mon-fri", hour=7, minute=55),
        id="freeze_and_run_brief",
        replace_existing=True,
    )
    scheduler.add_job(
        _purge_alias_maps,
        trigger=CronTrigger(hour=16, minute=0),
        id="purge_alias_maps",
        replace_existing=True,
    )
    scheduler.add_job(
        _retention_sweep,
        trigger=CronTrigger(hour=3, minute=30),
        id="retention_sweep",
        replace_existing=True,
    )
    scheduler.add_job(
        _aggregate_source_health,
        trigger=CronTrigger(minute="*/5"),
        id="source_health_snapshot",
        replace_existing=True,
    )
    scheduler.add_job(
        _research_worker_tick,
        trigger=CronTrigger(minute="*"),
        id="research_worker_tick",
        replace_existing=True,
    )
    scheduler.add_job(
        _refresh_symbol_maps,
        trigger=CronTrigger(day_of_week="mon", hour=4, minute=0),
        id="refresh_symbol_maps",
        replace_existing=True,
    )
    return scheduler
